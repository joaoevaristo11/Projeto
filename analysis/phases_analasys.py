import os

from openpyxl import Workbook, load_workbook
from openpyxl.utils import get_column_letter
from openpyxl.styles import Font
from openpyxl.chart import BarChart, Reference
from openpyxl.chart.label import DataLabelList


def ler_acao_txt(caminho_ficheiro):
    with open(caminho_ficheiro, 'r') as f:
        linhas = f.readlines()
        return [int(linha.strip()) for linha in linhas if linha.strip().isdigit()]

def calcular_percentagens_fases(dados, fases, comprimento):
    percentagens = {fase: [] for fase in fases}
    for agente, lista in dados.items():
        col_total = len(lista)
        distrib = {fase: lista.count(fase)/col_total*100 if col_total else 0 for fase in fases}
        for fase in fases:
            percentagens[fase].append(distrib[fase])
    return percentagens

def processar_testes_e_criar_excel(diretorio_base, nomes_testes):
    agentes = ['C0', 'C1', 'C2', 'C3', 'C4', 'C5', 'C6', 'C7', 'C8']
    fases = list(range(1, 10))
    todas_percentagens = []

    for nome_teste in nomes_testes:
        diretorio = os.path.join(diretorio_base, nome_teste)
        dados = {}

        for agente in agentes:
            caminho = os.path.join(diretorio, f"Agent Actions {agente}.txt")
            if not os.path.isfile(caminho):
                print(f":warning: Ficheiro não encontrado: {caminho}")
                return
            dados[agente] = ler_acao_txt(caminho)

        comprimento = max(len(v) for v in dados.values())
        percentagens = []

        for fase in fases:
            linha = []
            for agente in agentes:
                total = len(dados[agente])
                count = dados[agente].count(fase)
                perc = (count / total) * 100 if total else 0
                linha.append(perc)
            percentagens.append(linha)  # linha por fase
        todas_percentagens.append(percentagens)

    # Calcular médias
    medias = []
    for i in range(len(fases)):
        linha_media = []
        for j in range(len(agentes)):
            soma = sum(todas_percentagens[t][i][j] for t in range(len(nomes_testes)))
            media = soma / len(nomes_testes)
            linha_media.append(media)
        medias.append(linha_media)

    # Criar Excel
    caminho_saida = os.path.join(diretorio_base, 'medias_fases.xlsx')
    wb = Workbook()
    ws = wb.active
    ws.title = "Médias"

    # Cabeçalhos
    ws.cell(row=1, column=1, value="Fase").font = Font(bold=True)
    for i, agente in enumerate(agentes):
        ws.cell(row=1, column=2 + i, value=agente).font = Font(bold=True)

    # Inserir dados das médias
    for i, fase in enumerate(fases):
        ws.cell(row=2 + i, column=1, value=fase)
        for j in range(len(agentes)):
            ws.cell(row=2 + i, column=2 + j, value=medias[i][j])

    # Criar gráfico
    chart = BarChart()
    chart.type = "col"
    chart.style = 10
    chart.grouping = "clustered"
    chart.overlap = 0
    chart.title = "Média das Fases por Cruzamento"
    chart.y_axis.title = "Percentagem"
    chart.x_axis.title = "Fase"

    data = Reference(ws, min_col=2, max_col=6, min_row=1, max_row=10)
    cats = Reference(ws, min_col=1, min_row=2, max_row=10)

    chart.add_data(data, titles_from_data=True)
    chart.set_categories(cats)
    chart.dataLabels = DataLabelList()
    chart.dataLabels.showVal = True

    chart.height = 12
    chart.width = 24
    ws.add_chart(chart, "H2")

    wb.save(caminho_saida)
    print(f":white_check_mark: Excel com médias criado: {caminho_saida}")




if __name__ == "__main__":
    diretorio_base = r'E:\GGALVAO\BOLSA\Trabalhos\2_NN_2_cells_DQN\TLCS\models\model_8'
    nomes_testes = ["test_5000","test_10000","test_10003"]  # no meu caso tenho test_10000 a test_10004
    processar_testes_e_criar_excel(diretorio_base, nomes_testes)

