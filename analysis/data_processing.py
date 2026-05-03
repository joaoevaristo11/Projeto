import os
import numpy as np
from openpyxl import Workbook, load_workbook
from openpyxl.chart import LineChart, Reference
from openpyxl.utils import column_index_from_string
from openpyxl.chart.series import SeriesLabel

# CONFIGURAÇÕES
#models = {'Estratégia 4': r'./models/model_5'} # para ambiente real: ./models/real_env
models = {'Estratégia 4': r'./models/real_env'}
test_dirs = ['test_5000', 'test_10000', 'test_10003']
#excel_file = r'./analysis/valores_inteligente.xlsx'
excel_file = r'./analysis/valores_real.xlsx'

column_mapping = {
    'reward': ['A', 'B', 'C', 'D', 'E', 'F', 'G', 'H', 'I'],
    'queue': ['K', 'L', 'M', 'N', 'O', 'P', 'Q', 'R', 'S'],
    'ped_halting': ['U', 'V', 'W', 'X', 'Y', 'Z', 'AA', 'AB', 'AC'],
    'speed_med': ['AE', 'AF', 'AG', 'AH', 'AI', 'AJ', 'AK', 'AL', 'AM'],
    'avg_wt': ['AO', 'AP', 'AQ', 'AR', 'AS', 'AT', 'AU', 'AV', 'AW'],
    'avg_phase_time': ['AY', 'AZ', 'BA', 'BB', 'BC', 'BD', 'BE', 'BF', 'BG'],
    'lane_volume': ['BI']
}

reward_files      = [f'reward_C{i}.txt'                for i in range(1, 5)]  # C1..C4
queue_files       = [f'Queue_{i}.txt'                  for i in range(1, 5)]
ped_halting_files = [f'Pedestrian Halting C{i}.txt'    for i in range(1, 5)]
speed_med_files   = [f'Average Vehicle Speed C{i}.txt' for i in range(1, 5)]
wt_med_files      = [f'Average Waiting Time C{i}.txt'  for i in range(1, 5)]
pT_med_files      = [f'Average Phase Time in C{i}.txt' for i in range(1, 5)]
extra_files       = {f'Lane Volume.txt'}

# Inicializar Excel
wb = Workbook()
for sheet in models.keys():
    wb.create_sheet(title=sheet)
wb.remove(wb['Sheet'])
wb.save(excel_file)

def read_file_lines(path):
    with open(path, 'r') as f:
        return [float(line.strip()) for line in f if line.strip()]

def media_entre_testes(base_path, ficheiro_nome):
    valores = []
    for test_dir in test_dirs:
        path = os.path.join(base_path, test_dir, ficheiro_nome)
        linhas = read_file_lines(path)
        valores.append(linhas)
    return np.mean(np.array(valores), axis=0).tolist()

# Processar cada modelo
for sheet_name, model_path in models.items():
    print(f"🔄 A processar: {sheet_name}...")

    wb = load_workbook(excel_file)
    sheet = wb[sheet_name]

    # ── Cabeçalhos: índice começa em 1 (C1..C4) ──────────────────────────────
    for tipo, colunas in column_mapping.items():
        if isinstance(colunas, list):
            for i, col in enumerate(colunas):
                sheet[f'{col}1'] = f'{tipo}_{i+1}'   # FIX: era i, agora i+1
        else:
            sheet[f'{colunas}1'] = tipo

    # ── Reward ────────────────────────────────────────────────────────────────
    for idx, ficheiro in enumerate(reward_files):
        path = os.path.join(model_path, ficheiro)
        if not os.path.exists(path):
            continue
        valores = read_file_lines(path)
        for row_idx, valor in enumerate(valores, 2):
            col = column_mapping['reward'][idx]
            sheet[f'{col}{row_idx}'] = valor

    # ── Queue ─────────────────────────────────────────────────────────────────
    for idx, ficheiro in enumerate(queue_files):
        valores = media_entre_testes(model_path, ficheiro)
        for row_idx, valor in enumerate(valores, 2):
            col = column_mapping['queue'][idx]
            sheet[f'{col}{row_idx}'] = valor

    # ── Pedestrian Halting ────────────────────────────────────────────────────
    for idx, ficheiro in enumerate(ped_halting_files):
        valores = media_entre_testes(model_path, ficheiro)
        for row_idx, valor in enumerate(valores, 2):
            col = column_mapping['ped_halting'][idx]
            sheet[f'{col}{row_idx}'] = valor

    # ── Speed ─────────────────────────────────────────────────────────────────
    for idx, ficheiro in enumerate(speed_med_files):
        valores = media_entre_testes(model_path, ficheiro)
        for row_idx, valor in enumerate(valores, 2):
            col = column_mapping['speed_med'][idx]
            sheet[f'{col}{row_idx}'] = valor

    # ── Waiting Time ──────────────────────────────────────────────────────────
    for idx, ficheiro in enumerate(wt_med_files):
        valores = media_entre_testes(model_path, ficheiro)
        for row_idx, valor in enumerate(valores, 2):
            col = column_mapping['avg_wt'][idx]
            sheet[f'{col}{row_idx}'] = valor

    # ── Phase Time ────────────────────────────────────────────────────────────
    for idx, ficheiro in enumerate(pT_med_files):
        valores = media_entre_testes(model_path, ficheiro)
        for row_idx, valor in enumerate(valores, 2):
            col = column_mapping['avg_phase_time'][idx]
            sheet[f'{col}{row_idx}'] = valor

    # ── Lane Volume ───────────────────────────────────────────────────────────
    for idx, ficheiro in enumerate(extra_files):
        valores = media_entre_testes(model_path, ficheiro)
        for row_idx, valor in enumerate(valores, 2):
            col = column_mapping['lane_volume'][idx]
            sheet[f'{col}{row_idx}'] = valor

    wb.save(excel_file)

wb.save(excel_file)
print(f"\n✅ Valores calculados e guardados em:\n{excel_file}")